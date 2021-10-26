import os
from openpyxl import load_workbook, load_workbook

cwd = os.getcwd()
print('[*] Working in %s' % (cwd))

file = 'IPAM.xlsx'

wb = load_workbook(filename=file)

def delete_sheets_from_workbook(): # Create a fresh start with a single sheet
    allSheets = wb.sheetnames
    print("[-] Removing all existing sheets")
    for sheet in allSheets:
        del wb[sheet]

def create_ipam_sheets(): 
    sheets = ['IP Addresses', 'IP Ranges', 'Prefixes', 'Prefix and VLAN Roles', 'Aggregates', 'RIRs', 'VRFs', 'Route Targets', 'VLANs', 'VLAN Groups', 'Services']
    for sheet in sheets:
        sheet = wb.create_sheet(sheet)

def create_template_ip_addresses():
    sheet = wb['IP Addresses']
    options = ['address', 'vrf', 'tenant', 'status', 'role', 'device', 'virtual_machine', 'interface', 'is_primary','dns_name', 'description']
    for row in range(1):
        sheet.append(options)

def create_template_ip_ranges():
    sheet = wb['IP Ranges']
    options = ['start_address', 'end_address', 'vrf', 'tenant', 'status', 'role', 'description']
    for row in range(1):
        sheet.append(options)

def create_template_prefixes():
    sheet = wb['Prefixes']
    options = ['prefix', 'vrf', 'tenant', 'site', 'vlan_group', 'vlan', 'status', 'role', 'is_pool', 'mark_utilized', 'description']
    for row in range(1):
        sheet.append(options)

def create_template_prefixes_and_vlan_roles():
    sheet = wb['Prefix and VLAN Roles']
    options = ['name', 'slug', 'weight', 'description']
    for row in range(1):
        sheet.append(options)

def create_template_aggregates():
    sheet = wb['Aggregates']
    options = ['prefix', 'rir', 'tenant', 'date_added', 'description']
    for row in range(1):
        sheet.append(options)

def create_template_rirs():
    sheet = wb['RIRs']
    options = ['name', 'slug', 'is_private', 'description']
    for row in range(1):
        sheet.append(options)

def create_template_vrfs():
    sheet = wb['VRFs']
    options = ['name', 'rd', 'tenant', 'enforce_unique', 'description']
    for row in range(1):
        sheet.append(options)

def create_template_route_targets():
    sheet = wb['Route Targets']
    options = ['name', 'description', 'tenant']
    for row in range(1):
        sheet.append(options)

def create_template_vlans():
    sheet = wb['VLANs']
    options = ['site', 'group', 'vid', 'name', 'tenant', 'status', 'role', 'description']
    for row in range(1):
        sheet.append(options)

def create_template_vlan_groups():
    sheet = wb['VLAN Groups']
    options = ['name', 'slug', 'scope_type', 'scope_id', 'description']
    for row in range(1):
        sheet.append(options)

def create_template_services():
    sheet = wb['Services']
    options = ['device', 'virtual_name', 'name', 'protocol', 'ports', 'description', 'cf_format']
    for row in range(1):
        sheet.append(options)

def save_workbook():
    wb.save(file)
    print("[*] File %s saved in %s" % (file, cwd))

if __name__ == '__main__':
    try: 
        delete_sheets_from_workbook()
        create_ipam_sheets()
        create_template_ip_addresses()
        create_template_ip_ranges()
        create_template_prefixes()
        create_template_prefixes_and_vlan_roles()
        create_template_aggregates()
        create_template_rirs()
        create_template_vrfs()
        create_template_route_targets()
        create_template_vlans()
        create_template_vlan_groups()
        create_template_services()
        save_workbook()
    except:
        print("[!!] The action can't be completed because the file (%s) is open. Close the file and try again." % (file))
